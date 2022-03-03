import smtplib, ssl; from email.mime.text import MIMEText; from email.mime.multipart import MIMEMultipart
import openpyxl 
from datetime import datetime; import time

def main(): # Main func
    while(18 != datetime.now().hour or 0 != datetime.now().minute): # If entered time ! current time
        time.sleep(50) # Wait 50s if now ! required time
    birthday_check() # Next step 
    time.sleep(85800) # Wait 23 hours 50 minutes
    birthday_check()
    main() # Recursion

def birthday_check(): # Compares current date with contact's birthday
    current_month = datetime.now().month # Month shortcut
    current_day = datetime.now().day # Day shortcut
    
    person = [] 
    for r in range(1, ws.max_row+1): # For person
        for c in range(1, ws.max_column+1): # For attribute
            person.append(ws.cell(row = r, column = c).value) # Add cell to person
        if person[6] == current_month and person[7] == current_day: # If month and day == birth month and birth day
            compose_messages(person) # # Proceed to compose birthday messages 
        person = [] # Reset for next person

def compose_messages(birthday_person): # Builds messsages for send_message
    email = ws['J1'].value # User's email
    subject = '!!! BIRTHDAY ALERT !!!' # Email subject
    # Email text message
    text = """
        Hey! It is """+birthday_person[0]+"""'s birthday today, be sure to wish them a happy birthday!
        They're turning """+str((datetime.now().year-birthday_person[5])-1)+""".
        This is an automated message from contact_book.py
        """
    # Email HTML message
    html = """
        <html>
        <body>
            <p>Hey! It is """+birthday_person[0]+"""'s birthday today, be sure to wish them a happy birthday!<br>
            They're turning """+str((datetime.now().year-birthday_person[5])-1)+""".<br>
            This is an automated message from contact_book.py<br>
            </p>
        </body>
        </html>
        """
    smtp = 'smtp.'+ws['J1'].value.split('@')[1] # Email smtp
    send_message(email, subject, text, html, smtp) # Proceed to send message
    
    # Same logic as previous message
    email = birthday_person[3]
    subject = 'Happy Birthday!!!'
    text = """Happy birthday, """+birthday_person[0]+"""! Have a great """+str((datetime.now().year-birthday_person[5])-1)+""" birthday!        """
    html = """\
        <html>
        <body>
            <p>Happy birthday, """+birthday_person[0]+"""! Have a great """+str((datetime.now().year-birthday_person[5])-1)+""" birthday!<br>
            </p>
        </body>
        </html>
        """
    smtp = 'smtp.'+birthday_person[3].split('@')[1]
    send_message(email, subject, text, html, smtp)

def send_message(_email, _subject, _text, _html, _smtp):
    sender_email = "example@gmail.com " # Defines the sender's email address
    password = "password" # Sender's email address password

    message = MIMEMultipart("alternative") # Set multipart subtype
    message["Subject"] = (_subject) # Subject
    message["From"] = sender_email # Sender email address is determined
    message["To"] = _email # Receiver(s)'s email address(es) determined

    part1 = MIMEText(_text, "plain") # Builds MIME message
    part2 = MIMEText(_html, "html")
    message.attach(part1) 
    message.attach(part2)

    context = ssl.create_default_context() # Sends mail
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as server: 
        server.login(sender_email, password)
        server.sendmail(sender_email, _email, message.as_string(), 'high')

if __name__ == '__main__':
    # Excel setup
    wb = openpyxl.load_workbook('contact_book.xlsx')
    ws = wb['contact_book']
    main()
    
    